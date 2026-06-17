#!/usr/bin/env python3
"""
CDE Data Pipeline - California Department of Education Data Ingestion

This pipeline loads CDE downloadable data files (all 24 catalog domains) into
DuckDB / filesystem for downstream dbt transformations.

Features:
- Glob-based file discovery (no hardcoded filenames) across all CDE domains
- Three file types: tab-delimited .txt, multi-sheet .xlsx, caret-delimited .zip (SBAC)
- BOM stripping (\ufeff) on first column header
- Schema-drift tolerant: union by column name across years, missing cols -> None
- Suppression handling: ``*`` and empty strings -> None
- Per-domain dlt resources + selective --domain CLI loading
- Row-count logging per file
- Backward-compatible ``load_chronic_absenteeism`` entry point

See ``data/cde_raw/DATA_CATALOG.md`` for the full domain catalogue and quirks.
"""

import csv
import logging
import os
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional

import dlt
from dlt.common.pipeline import LoadInfo
from dlt.sources import DltResource

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Domain catalogue
# ---------------------------------------------------------------------------
# Each entry maps a domain key to its discovery + parsing configuration.
#   glob   : glob pattern relative to data_dir (all files are flat in data/raw/)
#   type   : "txt" (tab-delimited), "xlsx" (Excel), or "zip" (caret-delimited)
#   sheet  : substring matched against xlsx sheet names to find the data sheet
#            (None -> last sheet)
#   delim  : field delimiter for txt/zip ("\\t" default)
#   desc   : human-readable description
DOMAIN_CONFIG: Dict[str, Dict[str, Any]] = {
    "chronic_absenteeism": {
        "glob": "chronicabsenteeism*.txt",
        "type": "txt",
        "desc": "Chronic absenteeism rates (Style A)",
    },
    "absenteeism_reason": {
        "glob": "absenteeismreason*.txt",
        "type": "txt",
        "desc": "Absenteeism by reason (Style A)",
    },
    "cumulative_enrollment": {
        "glob": "cenroll*.txt",
        "type": "txt",
        "desc": "Cumulative enrollment (Style A, joined column names)",
    },
    "census_enrollment_rates": {
        "glob": "censusenrollratesdownload*.txt",
        "type": "txt",
        "desc": "Census enrollment rates (Style B)",
    },
    "chronic_absenteeism_dashboard": {
        "glob": "chronicdownload*.txt",
        "type": "txt",
        "desc": "Chronic absenteeism dashboard accountability (Style B)",
    },
    "suspension": {
        "glob": "suspension*.txt",
        "type": "txt",
        "desc": "Suspension counts/rates (Style A)",
    },
    "suspension_dashboard": {
        "glob": "suspdownload*.txt",
        "type": "txt",
        "desc": "Suspension dashboard accountability (Style B)",
    },
    "expulsion": {
        "glob": "expulsion*.txt",
        "type": "txt",
        "desc": "Expulsion counts/rates (Style A)",
    },
    "ela_dashboard": {
        "glob": "eladownload*.txt",
        "type": "txt",
        "desc": "ELA assessment dashboard (Style B)",
    },
    "elpac_dashboard": {
        "glob": "elpidownload*.txt",
        "type": "txt",
        "desc": "ELPAC English-learner proficiency dashboard (Style B)",
    },
    "sbac_caaspp": {
        "glob": "sb_ca*_all_csv_v1.zip",
        "type": "zip",
        "delim": "^",
        "desc": "SBAC/CAASPP assessment (caret-delimited inside zip)",
    },
    "homeless_enrollment": {
        "glob": "hse*.txt",
        "type": "txt",
        "desc": "Homeless student enrollment (Style A)",
    },
    "frpm": {
        "glob": "frpm*.xlsx",
        "type": "xlsx",
        "sheet": "School-Level",
        "desc": "Free/Reduced Price Meals (Excel, header at R2)",
    },
    "upc": {
        "glob": "cupc*.xlsx",
        "type": "xlsx",
        "sheet": "School-Level",
        "desc": "Unduplicated Pupil Count (Excel, header at R2)",
    },
    "restraint_seclusion": {
        "glob": "rsddata*.xlsx",
        "type": "xlsx",
        "sheet": None,
        "desc": "Restraint & seclusion (Excel, header at R2/R3)",
    },
    "school_directory": {
        "glob": "schldir.txt",
        "type": "txt",
        "desc": "School directory (Style C, PII)",
    },
    "cbedsora": {
        "glob": "cbedsora*.txt",
        "type": "txt",
        "desc": "CBEDS staff demographics, 'a'/'b' file pairs (Style C)",
    },
    "enrollment_by_grade": {
        "glob": "enrbygrade*.txt",
        "type": "txt",
        "desc": "Enrollment by grade (Style C)",
    },
    "enrollment_by_subgroup": {
        "glob": "enrbysubgrp*.txt",
        "type": "txt",
        "desc": "Enrollment by demographic subgroup (Style C)",
    },
    "class_assignment": {
        "glob": "classassign*.txt",
        "type": "txt",
        "desc": "Class assignment (Style C)",
    },
    "teacher_misassignment": {
        "glob": "teachermisassign*.txt",
        "type": "txt",
        "desc": "Teacher misassignment (Style C)",
    },
    "teacher_out_of_field": {
        "glob": "teacheroutoffield*.txt",
        "type": "txt",
        "desc": "Teacher out-of-field (Style C)",
    },
    "teacher_prep": {
        "glob": "teacherprep*.txt",
        "type": "txt",
        "desc": "Teacher preparation (Style C, 113 cols)",
    },
}

# Ordered list of domain keys (used for iteration + --domain=help)
ALL_DOMAINS: List[str] = list(DOMAIN_CONFIG.keys())


def _strip_bom(name: str) -> str:
    """Strip a leading UTF-8 BOM (\ufeff) from a column name."""
    if name and name.startswith("\ufeff"):
        return name[1:]
    return name


def _clean_value(value: Any) -> Any:
    """Normalise a raw cell value: ``*`` and empty strings -> None."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.replace("\r", "").strip()
        if v == "" or v == "*":
            return None
        return v
    return value


# ---------------------------------------------------------------------------
# Core loader
# ---------------------------------------------------------------------------
class CDEDataLoader:
    """Loader for California Department of Education data files."""

    def __init__(self, data_dir: str = ""):
        # Default to repo-relative "data/raw/" per the data catalogue.
        self.data_dir = data_dir or os.getenv("CDE_DATA_PATH", "data/raw")

    # -- generic file readers ---------------------------------------------

    def _read_tsv_file(
        self, file_path: str, delimiter: str = "\t"
    ) -> Iterator[Dict[str, Any]]:
        """Read a delimited text file and yield rows as dictionaries.

        Handles BOM on the first header name, ``\r`` in values, and suppression
        markers (``*`` / empty -> None).
        """
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                if reader.fieldnames:
                    # Strip BOM from the first column name in-place.
                    reader.fieldnames = [
                        _strip_bom(n) if i == 0 else n
                        for i, n in enumerate(reader.fieldnames)
                    ]
                for row in reader:
                    yield {k: _clean_value(v) for k, v in row.items()}
        except FileNotFoundError:
            logger.warning("File not found: %s", file_path)
            return
        except Exception as e:  # noqa: BLE001 - surface, don't crash the batch
            logger.error("Error reading %s: %s", file_path, e)
            return

    def _read_xlsx_file(
        self, file_path: str, sheet_substr: Optional[str] = None
    ) -> Iterator[Dict[str, Any]]:
        """Read an Excel data file, yielding rows as dictionaries.

        CDE Excel files put title/metadata rows above the real header on the
        data sheet. We locate the data sheet (by substring or last sheet) and
        auto-detect the header row (first row with >=3 populated cells),
        then emit the data rows that follow.
        """
        try:
            import openpyxl  # local import; heavy dependency
        except ImportError:  # pragma: no cover
            logger.error("openpyxl is required to read .xlsx CDE files")
            return

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            logger.error("Error opening %s: %s", file_path, e)
            return

        # Resolve the data sheet.
        ws = None
        if sheet_substr:
            for name in wb.sheetnames:
                if sheet_substr.lower() in name.lower():
                    ws = wb[name]
                    break
        if ws is None:
            ws = wb[wb.sheetnames[-1]]  # fall back to last sheet

        rows_iter = ws.iter_rows(values_only=True)

        # Auto-detect header row: first row (within the first ~5) that has >=3
        # non-None cells. This skips title blocks (1 cell) and the suppression
        # note row in rsddata2425.
        header: List[str] = []
        header_found = False
        for _ in range(5):
            try:
                row = next(rows_iter)
            except StopIteration:
                break
            populated = [c for c in row if c is not None]
            if not header_found and len(populated) >= 3:
                header = [_strip_bom(str(c)) if c is not None else f"col_{i}"
                           for i, c in enumerate(row)]
                header_found = True
                break
        if not header_found:
            logger.warning("No header row detected in %s", file_path)
            wb.close()
            return

        for row in rows_iter:
            if row is None or all(c is None for c in row):
                continue
            yield {header[i]: _clean_value(row[i])
                   for i in range(min(len(header), len(row)))}
        wb.close()

    def _read_zip_file(
        self, file_path: str, delimiter: str = "^"
    ) -> Iterator[Dict[str, Any]]:
        """Read a caret-delimited data file inside a .zip (SBAC/CAASPP).

        Picks the largest ``all_csv_v1.txt`` member when present, otherwise the
        first ``.txt`` member. Handles BOM + suppression markers.
        """
        try:
            with zipfile.ZipFile(file_path) as zf:
                members = [m for m in zf.namelist() if m.lower().endswith(".txt")]
                if not members:
                    logger.warning("No .txt member inside %s", file_path)
                    return
                # Prefer the full 'all_csv_v1' data member.
                target = next(
                    (m for m in members if "all_csv_v1" in m.lower()), members[0]
                )
                with zf.open(target) as inner:
                    reader = csv.DictReader(
                        (line.decode("utf-8", errors="replace") for line in inner),
                        delimiter=delimiter,
                    )
                    if reader.fieldnames:
                        reader.fieldnames = [
                            _strip_bom(n) if i == 0 else n
                            for i, n in enumerate(reader.fieldnames)
                        ]
                    for row in reader:
                        yield {k: _clean_value(v) for k, v in row.items()}
        except FileNotFoundError:
            logger.warning("File not found: %s", file_path)
            return
        except Exception as e:  # noqa: BLE001
            logger.error("Error reading zip %s: %s", file_path, e)
            return

    # -- file discovery ---------------------------------------------------

    def _discover_files(self, pattern: str) -> List[Path]:
        """Return sorted list of files matching ``pattern`` under data_dir."""
        base = Path(self.data_dir)
        if not base.exists():
            logger.warning("Data directory does not exist: %s", base)
            return []
        files = sorted(base.glob(pattern))
        # Exclude obvious non-data duplicates (e.g. 'schldir (1).txt').
        files = [f for f in files if "(" not in f.name]
        return files

    def _reader_for(self, file_path: Path, cfg: Dict[str, Any]) -> Iterator[Dict[str, Any]]:
        """Dispatch to the right reader based on file type."""
        ftype = cfg.get("type", "txt")
        if ftype == "txt":
            return self._read_tsv_file(str(file_path), delimiter=cfg.get("delim", "\t"))
        if ftype == "xlsx":
            return self._read_xlsx_file(str(file_path), sheet_substr=cfg.get("sheet"))
        if ftype == "zip":
            return self._read_zip_file(str(file_path), delimiter=cfg.get("delim", "^"))
        raise ValueError(f"Unknown file type '{ftype}' for {file_path}")

    # -- header scanning (for schema-drift union) -------------------------

    def _scan_txt_header(self, file_path: Path, delimiter: str) -> List[str]:
        """Read only the header line of a delimited text file."""
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                first_line = f.readline()
            names = first_line.rstrip("\n").rstrip("\r").split(delimiter)
            return [_strip_bom(n) if i == 0 else n for i, n in enumerate(names)]
        except Exception as e:  # noqa: BLE001
            logger.warning("Could not scan header of %s: %s", file_path, e)
            return []

    def _scan_xlsx_header(
        self, file_path: Path, sheet_substr: Optional[str] = None
    ) -> List[str]:
        """Read only the detected header row of an Excel data sheet."""
        try:
            import openpyxl
        except ImportError:
            return []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            logger.warning("Could not open %s: %s", file_path, e)
            return []
        ws = None
        if sheet_substr:
            for name in wb.sheetnames:
                if sheet_substr.lower() in name.lower():
                    ws = wb[name]
                    break
        if ws is None:
            ws = wb[wb.sheetnames[-1]]
        header: List[str] = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            populated = [c for c in row if c is not None]
            if len(populated) >= 3:
                header = [_strip_bom(str(c)) if c is not None else f"col_{i}"
                           for i, c in enumerate(row)]
                break
        wb.close()
        return header

    def _scan_zip_header(self, file_path: Path, delimiter: str) -> List[str]:
        """Read only the header line inside a .zip data file."""
        try:
            with zipfile.ZipFile(file_path) as zf:
                members = [m for m in zf.namelist() if m.lower().endswith(".txt")]
                if not members:
                    return []
                target = next(
                    (m for m in members if "all_csv_v1" in m.lower()), members[0]
                )
                with zf.open(target) as inner:
                    first_line = inner.readline().decode("utf-8", errors="replace")
            names = first_line.rstrip("\n").rstrip("\r").split(delimiter)
            return [_strip_bom(n) if i == 0 else n for i, n in enumerate(names)]
        except Exception as e:  # noqa: BLE001
            logger.warning("Could not scan header of %s: %s", file_path, e)
            return []

    def _scan_header(self, file_path: Path, cfg: Dict[str, Any]) -> List[str]:
        """Dispatch to the right header scanner based on file type."""
        ftype = cfg.get("type", "txt")
        if ftype == "txt":
            return self._scan_txt_header(file_path, cfg.get("delim", "\t"))
        if ftype == "xlsx":
            return self._scan_xlsx_header(file_path, sheet_substr=cfg.get("sheet"))
        if ftype == "zip":
            return self._scan_zip_header(file_path, cfg.get("delim", "^"))
        return []

    def _collect_columns(self, files: List[Path], cfg: Dict[str, Any]) -> List[str]:
        """First pass: scan headers of all files to build the column union."""
        all_columns: List[str] = []
        for f in files:
            for col in self._scan_header(f, cfg):
                if col and col not in all_columns:
                    all_columns.append(col)
        return all_columns

    # -- schema-drift-tolerant union loader -------------------------------

    def _load_domain(
        self, domain_key: str
    ) -> Iterator[Dict[str, Any]]:
        """Load all files for a domain, unioning columns across years.

        Two-pass approach for schema-drift tolerance:
        1. Scan headers of all matching files to build the full column union.
        2. Read data rows, filling missing columns with None so every emitted
           row has the same key set regardless of which year it came from.
        """
        cfg = DOMAIN_CONFIG[domain_key]
        files = self._discover_files(cfg["glob"])
        if not files:
            logger.info("No files matched %s for domain '%s'", cfg["glob"], domain_key)
            return

        # Pass 1: collect the union of all column names across years.
        all_columns = self._collect_columns(files, cfg)
        logger.info(
            "Domain '%s': %d files, %d union columns", domain_key, len(files), len(all_columns)
        )

        # Pass 2: emit data rows with the full column set.
        for file_path in files:
            logger.info("Loading %s ...", file_path.name)
            row_count = 0
            for row in self._reader_for(file_path, cfg):
                out = {col: row.get(col) for col in all_columns}
                out["_loaded_at"] = datetime.now().isoformat()
                out["_source_file"] = file_path.name
                out["_domain"] = domain_key
                yield out
                row_count += 1
            logger.info("  loaded %s rows from %s", f"{row_count:,}", file_path.name)

    # -- per-domain public loaders ---------------------------------------

    def load_chronic_absenteeism(
        self, academic_year: Optional[str] = None
    ) -> Iterator[Dict[str, Any]]:
        """Load chronic absenteeism data (backward compatible).

        Args:
            academic_year: 2-digit year suffix (e.g. "24" for 2023-24), or
                None to load all available years.
        Yields:
            Chronic absenteeism records with ``_loaded_at``/``_source_file``.
        """
        if academic_year:
            pattern = f"chronicabsenteeism{academic_year}*.txt"
        else:
            pattern = "chronicabsenteeism*.txt"

        files = self._discover_files(pattern)
        for file_path in files:
            logger.info("Loading %s ...", file_path.name)
            row_count = 0
            for row in self._read_tsv_file(str(file_path)):
                row["_loaded_at"] = datetime.now().isoformat()
                row["_source_file"] = file_path.name
                yield row
                row_count += 1
            logger.info("  loaded %s rows from %s", f"{row_count:,}", file_path.name)

    # Generic per-domain loaders (one method per catalogue domain).
    def load_absenteeism_reason(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("absenteeism_reason")

    def load_cumulative_enrollment(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("cumulative_enrollment")

    def load_census_enrollment_rates(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("census_enrollment_rates")

    def load_chronic_absenteeism_dashboard(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("chronic_absenteeism_dashboard")

    def load_suspension(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("suspension")

    def load_suspension_dashboard(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("suspension_dashboard")

    def load_expulsion(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("expulsion")

    def load_ela_dashboard(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("ela_dashboard")

    def load_elpac_dashboard(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("elpac_dashboard")

    def load_sbac_caaspp(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("sbac_caaspp")

    def load_homeless_enrollment(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("homeless_enrollment")

    def load_frpm(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("frpm")

    def load_upc(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("upc")

    def load_restraint_seclusion(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("restraint_seclusion")

    def load_school_directory(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("school_directory")

    def load_cbedsora(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("cbedsora")

    def load_enrollment_by_grade(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("enrollment_by_grade")

    def load_enrollment_by_subgroup(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("enrollment_by_subgroup")

    def load_class_assignment(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("class_assignment")

    def load_teacher_misassignment(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("teacher_misassignment")

    def load_teacher_out_of_field(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("teacher_out_of_field")

    def load_teacher_prep(self) -> Iterator[Dict[str, Any]]:
        yield from self._load_domain("teacher_prep")


# ---------------------------------------------------------------------------
# dlt source: one resource per domain
# ---------------------------------------------------------------------------
# Map domain key -> loader method name on CDEDataLoader.
_DOMAIN_METHODS: Dict[str, str] = {
    "chronic_absenteeism": "load_chronic_absenteeism",
    "absenteeism_reason": "load_absenteeism_reason",
    "cumulative_enrollment": "load_cumulative_enrollment",
    "census_enrollment_rates": "load_census_enrollment_rates",
    "chronic_absenteeism_dashboard": "load_chronic_absenteeism_dashboard",
    "suspension": "load_suspension",
    "suspension_dashboard": "load_suspension_dashboard",
    "expulsion": "load_expulsion",
    "ela_dashboard": "load_ela_dashboard",
    "elpac_dashboard": "load_elpac_dashboard",
    "sbac_caaspp": "load_sbac_caaspp",
    "homeless_enrollment": "load_homeless_enrollment",
    "frpm": "load_frpm",
    "upc": "load_upc",
    "restraint_seclusion": "load_restraint_seclusion",
    "school_directory": "load_school_directory",
    "cbedsora": "load_cbedsora",
    "enrollment_by_grade": "load_enrollment_by_grade",
    "enrollment_by_subgroup": "load_enrollment_by_subgroup",
    "class_assignment": "load_class_assignment",
    "teacher_misassignment": "load_teacher_misassignment",
    "teacher_out_of_field": "load_teacher_out_of_field",
    "teacher_prep": "load_teacher_prep",
}

# dlt table name per domain (cde_<domain>).
_DOMAIN_TABLES: Dict[str, str] = {
    key: f"cde_{key}" for key in DOMAIN_CONFIG
}


@dlt.source(name="cde")
def cde_source(
    data_dir: Optional[str] = None,
    academic_year: Optional[str] = None,
    domains: Optional[List[str]] = None,
) -> List[DltResource]:
    """dlt source for California Department of Education data.

    Args:
        data_dir: Path to the CDE raw data directory (default ``data/raw``).
        academic_year: 2-digit year suffix for chronic absenteeism (kept for
            backward compatibility); None loads all years.
        domains: Optional list of domain keys to load. None loads all domains.
    """
    loader = CDEDataLoader(data_dir=data_dir or "")
    selected = domains or ALL_DOMAINS

    resources: List[DltResource] = []

    def _make_resource(method, table_name, academic_year_filter):
        """Build a dlt resource wrapping a loader method (closure factory)."""

        if academic_year_filter is not None:
            @dlt.resource(name=table_name, write_disposition="replace")
            def _r() -> Iterator[Dict[str, Any]]:
                yield from method(academic_year=academic_year_filter)
        else:
            @dlt.resource(name=table_name, write_disposition="replace")
            def _r() -> Iterator[Dict[str, Any]]:
                yield from method()

        return _r

    for domain_key in selected:
        if domain_key not in DOMAIN_CONFIG:
            logger.warning("Unknown domain '%s' - skipping", domain_key)
            continue

        table_name = _DOMAIN_TABLES[domain_key]
        method_name = _DOMAIN_METHODS[domain_key]
        method = getattr(loader, method_name)
        # Only chronic_absenteeism supports the academic_year filter.
        ay_filter = academic_year if domain_key == "chronic_absenteeism" else None
        resources.append(_make_resource(method, table_name, ay_filter))

    return resources


# ---------------------------------------------------------------------------
# Pipeline runner
# ---------------------------------------------------------------------------
def run_cde_pipeline(
    destination_type: str = "duckdb",
    dataset_name: str = "cde_raw",
    data_dir: Optional[str] = None,
    academic_year: Optional[str] = None,
    domains: Optional[List[str]] = None,
) -> LoadInfo:
    """Run the CDE dlt pipeline.

    Args:
        destination_type: ``duckdb`` (default) or ``filesystem``.
        dataset_name: Name for the dataset.
        data_dir: Path to CDE raw data directory.
        academic_year: 2-digit year suffix for chronic absenteeism.
        domains: Optional list of domain keys to load (None = all).
    Returns:
        LoadInfo object with pipeline execution details.
    """
    if academic_year:
        logger.info("Loading CDE chronic absenteeism for year 20%s", academic_year)
    if domains:
        logger.info("Loading CDE domains: %s", ", ".join(domains))
    else:
        logger.info("Loading ALL available CDE domains (multi-year)")

    if destination_type == "duckdb":
        db_path = os.getenv("DUCKDB_DATABASE_PATH", "data/analytics.duckdb")
        db_path = os.path.abspath(db_path)
        pipeline = dlt.pipeline(
            pipeline_name="cde_to_duckdb",
            destination=dlt.destinations.duckdb(db_path),
            dataset_name=dataset_name,
        )
    else:
        stage1_path = os.getenv("STAGE1_PATH", "./oss_framework/data/stage1")
        pipeline = dlt.pipeline(
            pipeline_name="cde_to_stage1",
            destination=dlt.destinations.filesystem(
                bucket_url=f"{stage1_path}/transactional/cde"
            ),
            dataset_name=dataset_name,
        )

    source = cde_source(
        data_dir=data_dir, academic_year=academic_year, domains=domains
    )
    info = pipeline.run(source)

    logger.info("Pipeline completed successfully")
    logger.info("  Pipeline: %s", info.pipeline.pipeline_name)
    logger.info("  Destination: %s", destination_type)
    logger.info("  Dataset: %s", dataset_name)
    logger.info("  Loads: %s", len(info.loads_ids))
    return info


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _parse_cli_args(argv: List[str]) -> Dict[str, Any]:
    """Parse simple ``--key=value`` style CLI arguments."""
    opts: Dict[str, Any] = {
        "destination": "duckdb",
        "academic_year": None,
        "data_dir": None,
        "domains": None,
    }
    for arg in argv:
        if arg.startswith("--year="):
            opts["academic_year"] = arg.split("=", 1)[1]
        elif arg.startswith("--data-dir="):
            opts["data_dir"] = arg.split("=", 1)[1]
        elif arg.startswith("--domain="):
            value = arg.split("=", 1)[1]
            if value == "all" or value == "":
                opts["domains"] = None
            else:
                opts["domains"] = [d.strip() for d in value.split(",") if d.strip()]
        elif arg == "--filesystem":
            opts["destination"] = "filesystem"
        elif arg in ("--help", "-h"):
            opts["help"] = True
    return opts


if __name__ == "__main__":
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    args = _parse_cli_args(sys.argv[1:])

    if args.get("help"):
        print("CDE Data Pipeline")
        print("Usage: python cde_data_pipeline.py [options]")
        print("Options:")
        print("  --year=YY       Chronic absenteeism 2-digit year (e.g. 24)")
        print("  --data-dir=DIR  Path to CDE raw data (default: data/raw)")
        print("  --domain=LIST   Comma-separated domain keys (default: all)")
        print("  --filesystem    Write to filesystem (Parquet) instead of DuckDB")
        print("  --help          Show this help")
        print(f"\nAvailable domains ({len(ALL_DOMAINS)}):")
        for key in ALL_DOMAINS:
            print(f"  {key:32s} {DOMAIN_CONFIG[key]['desc']}")
        sys.exit(0)

    # Validate requested domains early.
    if args["domains"]:
        invalid = [d for d in args["domains"] if d not in DOMAIN_CONFIG]
        if invalid:
            print(f"❌ Unknown domain(s): {', '.join(invalid)}")
            print(f"Available: {', '.join(ALL_DOMAINS)}")
            sys.exit(1)

    run_cde_pipeline(
        destination_type=args["destination"],
        data_dir=args["data_dir"],
        academic_year=args["academic_year"],
        domains=args["domains"],
    )